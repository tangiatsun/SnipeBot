# Designed by Tangia Sun
# Fall 2022
# This is open source code for a Discord bot
#
#
# All player IDs are done using floats, since reading large values from xlsx
# files seem to cast it to floats which truncates the ID.
#
# I believe a Pandas implementation to something like a .csv file could fix
# the problem, but since the snipe_scores.xlsx is just for tracking, it should
# never be opened by a person. As long as its consistent, this shouldn't be a
# problem

#####

# Import modules
from distutils.util import change_root
import discord
import openpyxl
import os
import datetime

# List of participants
# {"id": None, "name": None "snipes": None, "sniped": None}
players = []

# Sheet we will be writing to
sheetNamePath = "./SnipeBotLogic/SnipeBot/snipe_scores.xlsx"

# Function to find file
def find(name, path):
    for root, dirs, files in os.walk(path):
        if name in files:
            return os.path.join(root, name)


# Writing the current players to a clean sheet
def writeSheet():
    print("Writing Sheet")
    # Make a new sheet
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet["A1"] = "Id"
    sheet["B1"] = "Name"
    sheet["C1"] = "Snipes"
    sheet["D1"] = "Sniped"
    sheet["F1"] = "Last Updated:"
    sheet["G1"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    # Save all the players in columns
    for i in players:
        player = (float(i["id"]), i["name"], i["snipes"], i["sniped"])
        sheet.append(player)
    wb.save(sheetNamePath)


# --- Extract scoresheet ---
# If scoresheet doesn't exist:
if find("snipe_scores.xlsx", "./SnipeBotLogic/SnipeBot") == None:
    # Creating empty scoresheet
    print("Making empty Scoresheet")
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet["A1"] = "Id"
    sheet["B1"] = "Name"
    sheet["C1"] = "Snipes"
    sheet["D1"] = "Sniped"
    sheet["F1"] = "Last Updated:"
    sheet["G1"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    wb.save(sheetNamePath)

# If scoresheet exists:
else:
    wb = openpyxl.load_workbook(sheetNamePath)
    sheet = wb.active

    # Valid scoresheet
    if (
        (sheet["A1"].value == "Id")
        & (sheet["B1"].value == "Name")
        & (sheet["C1"].value == "Snipes")
        & (sheet["D1"].value == "Sniped")
    ):
        print("Scoresheet Accepted")

        # Turn scoresheet into list items:
        for row in sheet.iter_rows(max_col=4, values_only=True):
            if (row[0] == "Id") or (row[0] == "") or (row[0] == None):
                continue
            else:
                participant = {
                    "id": float(row[0]),
                    "name": row[1],
                    "snipes": row[2],
                    "sniped": row[3],
                }
                players.append(participant)

    # Invalid scoresheet
    else:
        print("Scoresheet Rejected")
        # Creating scoresheet
        print("Making valid empty scoresheet")
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet["A1"] = "Id"
        sheet["B1"] = "Name"
        sheet["C1"] = "Snipes"
        sheet["D1"] = "Sniped"
        sheet["F1"] = "Last Updated:"
        sheet["G1"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        # Don't want to overwrite previous document, make new one
        sheetNamePath = "./SnipeBot/snipe_scores(1).xls"
        wb.save(sheetNamePath)


# Discord Code
intents = discord.Intents.default()
client = discord.Client(intents=intents)
# {"id": None, "snipes": 0, "sniped": 0}
print(players)


@client.event
async def on_ready():
    print(f"We have logged in as {client.user}")


@client.event
async def on_message(message):
    # Ignore message from bot
    if message.author == client.user:
        return
    if len(message.content) == 0:
        return

    # Takes the names of the sniper and the snipee(s)
    sniper = message.author
    sniperIn = False  # Check if sniper is an existing player

    # Filters through all persons sniped
    all_mentions = message.mentions
    for person in all_mentions:
        if person.discriminator == "4831":  # The bot
            continue
        else:
            # One or more people being sniped - a little redundant but speed is not important
            sniped = person
            snipedIn = False

            # Adding people to the list
            for d in players:  # Redudancy here, but checking for sniped
                if d["id"] == float(sniper.id):
                    sniperIn = True
                if d["id"] == float(sniped.id):
                    snipedIn = True
            if not snipedIn:  # Add sniped if not in
                players.append(
                    {
                        "id": float(sniped.id),
                        "name": sniped.name,
                        "snipes": 0,
                        "sniped": 0,
                    }
                )
            # Checking for attachments and setting up for reactions
            if len(message.attachments) > 0:
                print("-- New Snipe --")
                await message.reply(
                    sniped.name
                    + f" has been sniped by "
                    + sniper.name
                    + "! React to the picture with ✅ or ❎ to confirm or deny the snipe."
                )
                await message.add_reaction("✅")
                await message.add_reaction("❎")

            elif len(message.attachments) == 0:
                await message.channel.send(
                    "Please include a picture of the person you sniped."
                )
                return
            if sniper == sniped:
                await message.channel.send("You can't snipe yourself!")
                return

    # Adding in the sniper - only one person
    if not sniperIn:
        players.append(
            {"id": float(sniper.id), "name": sniper.name, "snipes": 0, "sniped": 0}
        )


@client.event
async def on_reaction_add(reaction, user):
    # Don't care about initial reactions
    if user == client.user:
        return
    print("-- Reaction Detected --")
    # Extract sniper and sniped for score and checking
    all_mentions = reaction.message.mentions
    sniper = reaction.message.author
    for person in all_mentions:
        if person.discriminator == "4831":  # The bot
            continue
        else:  # The person, or persons
            sniped = person
            # If the person that reacted is the person getting sniped
            if user == sniped:
                if reaction.emoji == "✅":  # Confirms the snipe
                    # Find in player list and adjust score
                    for d in players:
                        if d["id"] == float(sniped.id):
                            d["sniped"] += 1
                        elif d["id"] == float(sniper.id):
                            d["snipes"] += 1

    # Upgrade scoresheet
    writeSheet()


@client.event
async def on_raw_reaction_remove(payload):
    # user_id = person that removed message
    if payload.user_id == client.user.id:
        return
    print("-- Reaction Removed --")
    chan = client.get_channel(payload.channel_id)
    mess = await chan.fetch_message(payload.message_id)
    all_mentions = mess.mentions  # of type person
    sniper = mess.author
    for person in all_mentions:
        if person.discriminator == "4831":
            continue
        else:
            sniped = person

            if payload.user_id == sniped.id:
                if payload.emoji.name == "✅":
                    for d in players:
                        if d["id"] == float(sniped.id):
                            d["sniped"] -= 1
                        elif d["id"] == float(sniper.id):
                            d["snipes"] -= 1
    writeSheet()


# Insert your own bot token here
# client.run("")
